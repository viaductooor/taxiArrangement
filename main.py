import networkx as nx
import xlrd
import csv
import numpy as np
import xlwt
from entity import Passenger
from entity import Taxi
import csv
import openpyxl

HOME = r'E:\交通流量测算\taxiArrangement\files'
TRIPS = HOME + r'\trips_1.xls'
TRAVELTIME = HOME + r'\2010.2.2 8am new cost.xlsx'
REGION = HOME+r'\region_out.csv'
LOG = HOME+r'\log.xls'


def read_trips(url):
    map = {}
    table = xlrd.open_workbook(url).sheet_by_index(0)
    for i in range(1,table.nrows):
        t = xlrd.xldate_as_tuple(table.cell(i,7).value,0)
        min = t[4]
        row = table.row_values(i)
        no = row[0]
        loc = row[1]
        dest = row[2]
        passenger = Passenger(location=loc,destination=dest,waittime=min,number = no)
        # if min < 3:#TEST
        if map.get(min,-1) == -1:
            l = []
            l.append(passenger)
            map[min] = l
        else:
            map.get(min).append(passenger)
    return map

def read_traveltime(url):
    m = {}
    graph = nx.DiGraph()
    table = xlrd.open_workbook(url).sheet_by_index(0)
    for i in range(1,table.nrows):
        row = table.row_values(i)
        origin = int(row[0])
        dest = int(row[1])
        weight = float(row[2])
        graph.add_weighted_edges_from([(origin,dest,weight)])
        graph[origin][dest]['volume'] = 0
        graph[origin][dest]['total_volume']  = 0
    return graph

def read_region_map(url):
    m = {}
    i = 0
    reader = csv.reader(open(url,'r'))
    for row in reader:
        if i>0:
            node = int(row[0])
            region  = int(row[3])
            m[node] = region
        i += 1
    return m

def true_trip(graph,origin,dest):
    if graph.has_node(origin) and graph.has_node(dest):
        if nx.has_path(graph, origin,dest) and origin != dest:
            return True
    return False

def similarity(proute,troute,regionmap):
    if len(troute) == 0:
        #if the taxi has no passenger, the similarity should be assigned the highest value
        return 0.7
    else:
        mat1 = np.zeros((100,100))
        mat2 = np.zeros((100,100))
        for i in range(len(proute)-1):
            rstart = regionmap.get(proute[i])
            rend = regionmap.get(proute[i+1])
            mat1[rstart][rend] = 1
        for i in range(len(troute)-1):
            rstart = regionmap.get(troute[i])
            rend = regionmap.get(troute[i+1])
            mat2[rstart][rend] = 1
        for i in range(100):
            mat1[i][i] = 0
            mat2[i][i] = 0
        sum1 = np.sum(mat1)
        sum2 = np.sum(mat2)
        multisum = np.sum(mat1*mat2)
        return 0.5*(multisum/sum1+multisum/sum2)



def init_xlsx(g,wb):
    ws = wb.active
    i = 1
    for e in g.edges():
        begin,end = e[0],e[1]
        ws.cell(row=i,column=1).value = begin
        ws.cell(row=i,column=2).value = end
        i += 1

def carpool(min_similarity = 0.5):
    trips = read_trips(TRIPS)
    graph = read_traveltime(TRAVELTIME)
    region_map = read_region_map(REGION)
    taxi_list = []

    #initiate the workbook of volumes with carpool
    workbook_carpool = xlwt.Workbook()
    sheet_carpool = workbook_carpool.add_sheet('carpool')
    cols_carpool = ['init','end','with_carpool']
    for i in range(0,len(cols_carpool)):
        sheet_carpool.write(0,i,cols_carpool[i])
    volume_head_i = 1
    for edge in graph.edges():
        init = edge[0]
        end = edge[1]
        sheet_carpool.write(volume_head_i, 0, init)
        sheet_carpool.write(volume_head_i, 1, end)
        volume_head_i += 1

    #at the beginning, arrange every passenger a new taxi
    time = {'minute':0,'second':0}
    for p in trips.get(time['minute']):
        origin = p.location
        dest = p.destination
        if true_trip(graph,origin,dest):
            taxi = Taxi(graph,3,p.location)
            taxi.add_passenger(p)
            taxi_list.append(taxi)
    print(Passenger.on_taxi_num)

    #log
    xlsx_wb = openpyxl.Workbook()
    ws = xlsx_wb.active
    init_xlsx(graph,xlsx_wb)
    col = 3

    while Passenger.on_taxi_num > 0:
        time['second']+=1
        if time['second'] == 60:
            time['second'] = 0
            time['minute'] += 1

        #log
        print('time',time,'passenger_num',Passenger.on_taxi_num,'taxi_num',len(taxi_list))
        row = 1
        for e in graph.edges():
            begin, end = e[0], e[1]
            vol = graph[begin][end]['volume']
            ws.cell(row=row, column=col).value = vol
            row += 1
        row = 1
        col += 1

        for taxi in taxi_list:
            taxi.forward()

        #one minute passed, and new passengers need to take taxis
        if time['second'] == 0:
            if trips.get(time['minute'])!=None:
                for p in trips.get(time['minute']):
                    origin = p.location
                    dest = p.destination
                    if true_trip(graph,origin,dest):
                        proute = nx.shortest_path(graph,p.location,p.destination)
                        max_sim = 0
                        selected_taxi = None
                        #similarity
                        for t in taxi_list:
                            if region_map.get(t.location) == region_map.get(p.location) and len(t.passenger_list) < t.seat_num and nx.has_path(graph,t.location,p.location):
                                sim = similarity(proute,t.route,region_map)
                                if sim > max_sim:
                                    max_sim = sim
                                    selected_taxi = t
                        if max_sim > min_similarity:
                            try:
                                selected_taxi.add_passenger(p)
                            except nx.NetworkXNoPath:
                                taxi = Taxi(graph, 3, p.location)
                                taxi.add_passenger(p)
                                taxi_list.append(taxi)
                        else:
                            taxi = Taxi(graph,3,p.location)
                            taxi.add_passenger(p)
                            taxi_list.append(taxi)

    #write volume info into workbook after the whole run
    volume_head_i = 1
    for edge in graph.edges():
        totalvolume = graph[edge[0]][edge[1]]['total_volume']
        sheet_carpool.write(volume_head_i, 2, totalvolume)
        volume_head_i += 1
    workbook_carpool.save(HOME+r'\totalvolume_carpool.xls')
    xlsx_wb.save(HOME+r'\xlsx_flow.xlsx')

def original():
    trips = read_trips(TRIPS)
    graph = read_traveltime(TRAVELTIME)
    region_map = read_region_map(REGION)
    taxi_list = []

    # initiate the workbook of volumes with carpool
    workbook_carpool = xlwt.Workbook()
    sheet_carpool = workbook_carpool.add_sheet('carpool')
    cols_carpool = ['init', 'end', 'original']
    for i in range(0, len(cols_carpool)):
        sheet_carpool.write(0, i, cols_carpool[i])
    volume_head_i = 1
    for edge in graph.edges():
        init = edge[0]
        end = edge[1]
        sheet_carpool.write(volume_head_i, 0, init)
        sheet_carpool.write(volume_head_i, 1, end)
        volume_head_i += 1

    # at the beginning, arrange every passenger a new taxi
    time = {'minute': 0, 'second': 0}
    for p in trips.get(time['minute']):
        origin = p.location
        dest = p.destination
        if true_trip(graph, origin, dest):
            taxi = Taxi(graph, 3, p.location)
            taxi.add_passenger(p)
            taxi_list.append(taxi)

    while Passenger.on_taxi_num > 0:
        time['second'] += 1
        if time['second'] == 60:
            time['second'] = 0
            time['minute'] += 1

        print('time', time, 'passenger_num', Passenger.on_taxi_num, 'taxi_num', len(taxi_list))

        for taxi in taxi_list:
            taxi.forward()

        # one minute passed, and new passengers need to take taxis
        if time['second'] == 0:
            if trips.get(time['minute']) != None:
                for p in trips.get(time['minute']):
                    origin = p.location
                    dest = p.destination
                    if true_trip(graph, origin, dest):
                        taxi = Taxi(graph, 3, p.location)
                        taxi.add_passenger(p)
                        taxi_list.append(taxi)

    # write volume info into workbook after the whole run
    volume_head_i = 1
    for edge in graph.edges():
        totalvolume = graph[edge[0]][edge[1]]['total_volume']
        sheet_carpool.write(volume_head_i, 2, totalvolume)
        volume_head_i += 1
    workbook_carpool.save(HOME + r'\totalvolume_original.xls')

carpool()

